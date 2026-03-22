#!/usr/bin/env python3
"""Convert the 'Played' sheet from some_quartets_by_key.xlsx → played.json.

Filters out:
  - Rows marked N/A in the 'include?' column
  - Rows missing 'Equiv number of sharps' or 'Spelling of tonal center'

Usage: python played_to_json.py [input.xlsx] [output.json]
Defaults: some_quartets_by_key.xlsx → played.json
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


def clean(val):
    return str(val or "").strip()


def clean_piece(raw):
    """Stringify and strip Excel date/float artefacts."""
    s = str(raw or "").strip()
    if re.match(r"\d{4}-\d{2}-\d{2} 00:00:00$", s):
        return ""
    # Strip trailing ".0" from Excel float coercion (e.g. "42.0" → "42")
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def normalize_key(raw):
    """Normalize key spelling to match '>48 by key' sheet conventions.

    E.g. 'E-Flat' → 'E-flat', 'F-Sharp' → 'F-sharp', 'D-Sharp' → 'D-sharp'.
    """
    parts = raw.split("-", 1)
    if len(parts) == 2:
        return parts[0] + "-" + parts[1].lower()
    return raw


def convert(src: Path, dst: Path) -> None:
    wb = load_workbook(src)
    ws = wb["Played "]

    played = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        if all(v is None for v in vals):
            continue

        include = clean(vals[4]).upper()
        if include == "N/A":
            continue

        sharps = vals[5]
        tonal_center = clean(vals[6])
        if sharps is None or not tonal_center:
            continue

        date_val = vals[0]
        date_str = date_val.strftime("%Y-%m-%d") if date_val else ""

        played.append({
            "date":     date_str,
            "composer": clean(vals[1]),
            "piece":    clean_piece(vals[2]),
            "part":     clean(vals[3]),
            "include":  clean(vals[4]),
            "sharps":   float(sharps),
            "key":      normalize_key(tonal_center),
            "mode":     clean(vals[7]),
        })

    dst.write_text(json.dumps(played, indent=2, ensure_ascii=False))
    print(f"Wrote {len(played)} played pieces → {dst}")


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("some_quartets_by_key.xlsx")
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("played.json")
    convert(src, dst)
