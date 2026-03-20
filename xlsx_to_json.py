#!/usr/bin/env python3
"""Convert some_quartets_by_key.xlsx → quartets.json.

Usage: python xlsx_to_json.py [input.xlsx] [output.json]
Defaults: some_quartets_by_key.xlsx → quartets.json
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

COLOR_STATUS = {
    "00000000": "candidate",
    "FFFFF2CC": "uncertain",
    "FFCCCCCC": "alternate",
    "FF999999": "rejected",
    "FFF4CCCC": None,   # empty placeholder rows — skip
}

def cell_fill(cell):
    try:
        return cell.fill.fgColor.rgb
    except Exception:
        return "00000000"

def extract_yt_url(raw: str) -> str:
    """Pull the first YouTube URL out of a cell that may have surrounding text."""
    m = re.search(r"https?://[^\s]*youtu[^\s]*", raw or "")
    return m.group(0) if m else ""

def clean_piece(raw) -> str:
    """Stringify and strip Excel date artefacts like '2026-01-03 00:00:00'."""
    s = str(raw or "").strip()
    # Excel sometimes stores text dates as datetime objects
    if re.match(r"\d{4}-\d{2}-\d{2} 00:00:00$", s):
        return ""   # caller can decide; we just blank it
    return s

def convert(src: Path, dst: Path) -> None:
    wb = load_workbook(src)
    ws = wb[">48 by key"]

    pieces = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        # Column 0 is "Date played"; data columns start at index 1.
        composer = str(vals[5] or "").strip()
        if not composer:
            continue

        fill = cell_fill(row[0])
        status = COLOR_STATUS.get(fill, "candidate")
        if status is None:
            continue

        try:
            sharps = float(vals[1]) if vals[1] is not None else 0.0
        except (TypeError, ValueError):
            sharps = 0.0

        pieces.append({
            "sharps":   sharps,
            "key":      str(vals[2] or "").strip().rstrip("*?"),
            "mode":     str(vals[3] or "").strip(),
            "composer": composer,
            "piece":    clean_piece(vals[6]),
            "notes":    str(vals[7] or "").strip(),
            "scoreUrl": str(vals[9] or "").strip(),
            "ytUrl":    extract_yt_url(str(vals[10] or "")),
            "status":   status,
        })

    dst.write_text(json.dumps(pieces, indent=2, ensure_ascii=False))
    print(f"Wrote {len(pieces)} pieces → {dst}")

if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("some_quartets_by_key.xlsx")
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("quartets.json")
    convert(src, dst)
